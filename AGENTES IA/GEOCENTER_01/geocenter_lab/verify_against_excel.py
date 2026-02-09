import openpyxl
import sys
import json
from core.granulometry import calculate_granulometry
from core.limits import calculate_limits
# from core.classification import calculate_sucs

sys.stdout.reconfigure(encoding='utf-8')
file_path = r'c:\Users\Hp\Desktop\GEOCENTER_01\ENSAYOS.xlsx'

print("--- GEOCENTER LAB VERIFICATION ---\n")
print("Reading Excel Data...")

wb = openpyxl.load_workbook(file_path, data_only=True) # Read VALUES, not formulas

# 1. EXTRACT DATA
# ---------------

# Granulometry (Sheet Index 1)
sheet_gran = wb[wb.sheetnames[1]]
total_weight = float(sheet_gran['D17'].value)
print(f"Granulometry Total Weight: {total_weight}")

# Rows 24 to 40 roughly for sieves
sieves_data = []
# Map specific rows to standard sieves if possible, or just read list
# Based on previous dump: 
# Row 24: 3", Row 26: 3/4", Row 28: #4, Row 30: #16... (Values were interleaved?)
# Let's iterate and grab non-empty F cols
gran_inputs = []
# Specific mapping from previous analysis
# R24: 3", R25: 1 1/2"?? No, previous dump showed:
# R24: 3", R26: 3/4", R28: #4... wait, rows were 24, 25(1 1/2), 26(3/4).
# Previous dump:
# R24: 3" (0)
# R25: 1 1/2" (0)
# R26: 3/4" (468.3)
# R27: 3/8" (342.1)
# R28: #4 (245.1)
# ...
row_map = [
    (24, '3"', 75.0),
    (25, '1 1/2"', 37.5),
    (26, '3/4"', 19.0),
    (27, '3/8"', 9.5),
    (28, '# 4', 4.75),
    (29, '# 8', 2.36),   # Arema Gruesa
    (30, '# 16', 1.18),  # Arena Media
    (31, '# 30', 0.60),  # Not sure if 30 or similar
    (32, '# 50', 0.30),
    (33, '# 100', 0.15),
    (34, '# 200', 0.075)
]

for r_idx, label, size in row_map:
    w = sheet_gran[f'F{r_idx}'].value
    if w is None: w = 0
    gran_inputs.append({'size_label': label, 'opening_mm': size, 'weight_retained': float(w)})

# Limits (Sheet Index 6)
sheet_lim = wb[wb.sheetnames[6]]
# LL: Cols F, G, H, I
ll_cols = ['F', 'G', 'H', 'I']
ll_data = []
for col in ll_cols:
    blows = sheet_lim[f'{col}23'].value
    wt = sheet_lim[f'{col}24'].value
    dt = sheet_lim[f'{col}25'].value
    t = sheet_lim[f'{col}26'].value
    
    if blows and wt and dt and t:
        ll_data.append({
            'blows': float(blows), 'wet_tare': float(wt), 'dry_tare': float(dt), 'tare': float(t)
        })

# PL: Cols G, H, I (Rows 53, 54, 55)
pl_cols = ['G', 'H', 'I']
pl_data = []
for col in pl_cols:
    wt = sheet_lim[f'{col}53'].value
    dt = sheet_lim[f'{col}54'].value
    t = sheet_lim[f'{col}55'].value
    if wt and dt and t:
        pl_data.append({
            'wet_tare': float(wt), 'dry_tare': float(dt), 'tare': float(t)
        })

# 2. RUN CALCULATIONS
# -------------------

print(f"\nRunning Granulometry Check (Inputs: {len(gran_inputs)} sieves)...")
gran_res = calculate_granulometry({'total_dry_weight': total_weight, 'sieves': gran_inputs})

print(f"Running Limits Check (LL Pts: {len(ll_data)}, PL Pts: {len(pl_data)})...")
lim_res = calculate_limits({'ll_data': ll_data, 'pl_data': pl_data})


# 3. COMPARE RESULTS
# ------------------

# Granulometry Comparison
print("\n--- GRANULOMETRY COMPARISON ---")
print(f"{'Tamiz':<10} | {'Excel %Pasa':<12} | {'Software %Pasa':<14} | {'Diff':<6}")
print("-" * 50)
for i, item in enumerate(row_map):
    r_idx, label, _ = item
    # Excel Calculation Cell: Column I (Passing) usually?
    # Previous dump: R24 Col I formula = 100 - H24 (Cumulative). 
    # Let's read Value from Col I
    excel_val = sheet_gran[f'I{r_idx}'].value
    if excel_val is None: excel_val = 0
    excel_val = round(float(excel_val), 2)
    
    soft_val = gran_res['data'][i]['percent_passing']
    
    diff = round(abs(excel_val - soft_val), 2)
    status = "OK" if diff < 0.1 else "DIFF"
    print(f"{label:<10} | {excel_val:<12} | {soft_val:<14} | {status}")


# Limits Comparison
print("\n--- LIMITS COMPARISON ---")
# Excel Results: LL in N20?? No, scanning showed LL in Row 60 Col D formula =+N37?
# Let's verify where LL/PL result is. Row 60.
# Row 60: LL (Col D?), PL (Col G?), IP (Col I?) -> previous dump was confusing on headers vs cells
# Row 60: ['', 'Límite Liquido (L.L.) = ', '', '=+N37', 'Límite Plástico (L.P.) = ', '', '=+AVERAGE(G58:I58)']
# Actually LL result seems to be in cell D60? No, list showed '=+N37' is at index 3 which is D column.
# Let's read the values at D60, G60, I60 ?? Or read N37?
# N37 might be where the LL calculation was done. 
# Let's just trust the software calculation and print it vs the N37 value if we can read it.
excel_ll = sheet_lim['N37'].value 
# If N37 is far away, maybe just check D60 (which links to N37)
if excel_ll is None: excel_ll = sheet_lim['D60'].value

# PL Average
# Formula was AVERAGE(G58:I58)
g58 = sheet_lim['G58'].value
h58 = sheet_lim['H58'].value
i58 = sheet_lim['I58'].value
excel_pl = (g58+h58+i58)/3 if (g58 and h58 and i58) else 0

print(f"Limit      | Excel Value  | Software Val | Diff")
print("-" * 50)

soft_ll = lim_res['results']['LL']
print(f"{'LL':<10} | {str(excel_ll):<12} | {soft_ll:<12} | {round(abs((float(excel_ll or 0))-soft_ll), 2)}")

soft_pl = lim_res['results']['PL']
print(f"{'PL':<10} | {round(excel_pl, 2):<12} | {soft_pl:<12} | {round(abs(excel_pl-soft_pl), 2)}")

soft_pi = lim_res['results']['PI']
excel_pi = (float(excel_ll or 0) - excel_pl)
print(f"{'PI':<10} | {round(excel_pi, 2):<12} | {soft_pi:<12} | {round(abs(excel_pi-soft_pi), 2)}")

print("\nVerification Finished.")
